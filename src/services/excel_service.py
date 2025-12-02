"""
Excel Service - Import/Export operations for Excel files
Provides import with preview and validation, export with formatting preservation,
duplicate handling, and progress indicators for large files

Requirements: 11.1, 11.2, 11.3, 11.4, 11.5, 17.1, 17.3, 17.4
"""
import logging
from typing import List, Dict, Optional, Any, Callable
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.database.enhanced_db_manager import EnhancedDatabaseManager
from src.models.trip import Trip
from src.utils.error_handler import ImportExportError, ValidationError, handle_errors, ErrorHandler


logger = logging.getLogger(__name__)


class DuplicateHandling:
    """Enum for duplicate handling strategies"""
    SKIP = "skip"
    OVERWRITE = "overwrite"
    CREATE_NEW = "create_new"


class ExcelService:
    """
    Service for Excel import/export operations with validation and formatting.
    
    Features:
    - Import with preview and validation
    - Export with formatting preservation
    - Duplicate handling: skip, overwrite, create new
    - Progress indicators for large files
    - Error reporting with line numbers
    """
    
    def __init__(self, db_manager: EnhancedDatabaseManager):
        """
        Initialize Excel Service
        
        Args:
            db_manager: Database manager instance
        """
        self.db = db_manager
    
    # ========================================================================
    # Import Operations
    # ========================================================================
    
    def preview_excel_file(self, file_path: str, max_rows: int = 10) -> Dict[str, Any]:
        """
        Preview Excel file before import
        
        Args:
            file_path: Path to Excel file
            max_rows: Maximum number of rows to preview
        
        Returns:
            Dictionary containing:
            - columns: List of column names
            - preview_data: List of preview rows (max_rows)
            - total_rows: Total number of rows in file
            - validation_errors: List of validation errors found in preview
            
        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If file format is invalid
        """
        try:
            # Check if file exists
            if not Path(file_path).exists():
                raise FileNotFoundError(f"File không tồn tại: {file_path}")
            
            # Read Excel file
            df = pd.read_excel(file_path)
            
            # Get column names
            columns = df.columns.tolist()
            
            # Get preview data
            preview_rows = min(max_rows, len(df))
            preview_data = df.head(preview_rows).to_dict('records')
            
            # Validate preview data
            validation_errors = []
            for idx, row in enumerate(preview_data, start=2):  # Start from row 2 (after header)
                errors = self._validate_row(row, idx)
                if errors:
                    validation_errors.extend(errors)
            
            logger.info(f"Previewed {preview_rows} rows from {file_path}")
            
            return {
                'columns': columns,
                'preview_data': preview_data,
                'total_rows': len(df),
                'validation_errors': validation_errors
            }
            
        except FileNotFoundError:
            raise
        except Exception as e:
            logger.error(f"Error previewing Excel file: {e}")
            raise ValueError(f"Không thể đọc file Excel: {str(e)}")
    
    def import_excel_file(
        self,
        file_path: str,
        duplicate_handling: str = DuplicateHandling.SKIP,
        progress_callback: Optional[Callable[[int, int, str], None]] = None
    ) -> Dict[str, Any]:
        """
        Import data from Excel file with validation and duplicate handling
        
        Args:
            file_path: Path to Excel file
            duplicate_handling: How to handle duplicates (skip, overwrite, create_new)
            progress_callback: Optional callback function(current, total, message)
        
        Returns:
            Dictionary containing:
            - success_count: Number of successfully imported records
            - skipped_count: Number of skipped records
            - error_count: Number of failed records
            - errors: List of error details with line numbers
            - imported_trips: List of imported Trip objects
            
        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If file format is invalid or duplicate_handling is invalid
        """
        try:
            # Validate duplicate handling strategy
            valid_strategies = [DuplicateHandling.SKIP, DuplicateHandling.OVERWRITE, DuplicateHandling.CREATE_NEW]
            if duplicate_handling not in valid_strategies:
                raise ValueError(f"Invalid duplicate handling: {duplicate_handling}")
            
            # Check if file exists
            if not Path(file_path).exists():
                raise FileNotFoundError(f"File không tồn tại: {file_path}")
            
            # Read Excel file
            df = pd.read_excel(file_path)
            total_rows = len(df)
            
            # Initialize counters
            success_count = 0
            skipped_count = 0
            error_count = 0
            errors = []
            imported_trips = []
            
            # Process each row
            for idx, row in df.iterrows():
                row_number = idx + 2  # Excel row number (1-based + header)
                
                # Update progress
                if progress_callback:
                    progress_callback(idx + 1, total_rows, f"Processing row {row_number}...")
                
                try:
                    # Convert row to dictionary and clean NaN values
                    row_dict = row.to_dict()
                    row_dict = {k: (v if pd.notna(v) else None) for k, v in row_dict.items()}
                    
                    # Validate row
                    validation_errors = self._validate_row(row_dict, row_number)
                    if validation_errors:
                        errors.extend(validation_errors)
                        error_count += 1
                        continue
                    
                    # Prepare trip data
                    trip_data = self._prepare_trip_data(row_dict)
                    
                    # Handle duplicates
                    result = self._handle_duplicate(trip_data, duplicate_handling)
                    
                    if result['action'] == 'skip':
                        skipped_count += 1
                        logger.debug(f"Skipped duplicate at row {row_number}: {trip_data.get('ma_chuyen')}")
                    elif result['action'] == 'import':
                        trip = result['trip']
                        imported_trips.append(trip)
                        success_count += 1
                        logger.debug(f"Imported row {row_number}: {trip.ma_chuyen}")
                    
                except Exception as e:
                    error_count += 1
                    error_msg = f"Row {row_number}: {str(e)}"
                    errors.append(error_msg)
                    logger.error(error_msg)
            
            # Final progress update
            if progress_callback:
                progress_callback(total_rows, total_rows, "Import completed")
            
            logger.info(f"Import completed: {success_count} success, {skipped_count} skipped, {error_count} errors")
            
            return {
                'success_count': success_count,
                'skipped_count': skipped_count,
                'error_count': error_count,
                'errors': errors,
                'imported_trips': imported_trips
            }
            
        except FileNotFoundError:
            raise
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"Error importing Excel file: {e}")
            raise
    
    def _validate_row(self, row_dict: Dict[str, Any], row_number: int) -> List[str]:
        """
        Validate a single row of data
        
        Args:
            row_dict: Row data as dictionary
            row_number: Row number for error reporting
        
        Returns:
            List of validation error messages
        """
        errors = []
        
        # Helper function to get value from alternative column names
        def get_value(possible_names):
            for name in possible_names:
                if name in row_dict:
                    return row_dict[name]
            return None
        
        # Check required fields with alternative names
        khach_hang = get_value(['khach_hang', 'Khách hàng', 'Khach hang', 'Customer'])
        if not khach_hang:
            errors.append(f"Row {row_number}: Khách hàng không được để trống")
        
        gia_ca = get_value(['gia_ca', 'Giá cả', 'Gia ca', 'Price'])
        if gia_ca is None:
            errors.append(f"Row {row_number}: Giá cả không được để trống")
        
        # Validate numeric fields with alternative names
        numeric_field_mappings = {
            'gia_ca': ['gia_ca', 'Giá cả', 'Gia ca', 'Price'],
            'khoan_luong': ['khoan_luong', 'Khoán lương', 'Khoan luong', 'Salary'],
            'chi_phi_khac': ['chi_phi_khac', 'Chi phí khác', 'Chi phi khac', 'Other Costs']
        }
        
        for field, possible_names in numeric_field_mappings.items():
            value = get_value(possible_names)
            if value is not None:
                try:
                    num_value = int(value) if isinstance(value, (int, float)) else int(str(value))
                    if num_value < 0:
                        errors.append(f"Row {row_number}: {field} không được âm")
                except (ValueError, TypeError):
                    errors.append(f"Row {row_number}: {field} phải là số")
        
        # Validate trip code format if provided
        ma_chuyen = get_value(['ma_chuyen', 'Mã chuyến', 'Ma chuyen', 'Trip Code'])
        if ma_chuyen:
            import re
            if not re.match(r'^C\d+$', str(ma_chuyen)):
                errors.append(f"Row {row_number}: Mã chuyến phải có định dạng C theo sau bởi số (ví dụ: C001)")
        
        return errors
    
    def _prepare_trip_data(self, row_dict: Dict[str, Any]) -> Dict[str, Any]:
        """
        Prepare trip data from row dictionary
        
        Args:
            row_dict: Row data as dictionary
        
        Returns:
            Prepared trip data dictionary
        """
        # Map column names to model fields (handle various column name formats)
        field_mapping = {
            'ma_chuyen': ['ma_chuyen', 'Mã chuyến', 'Ma chuyen', 'Trip Code'],
            'khach_hang': ['khach_hang', 'Khách hàng', 'Khach hang', 'Customer'],
            'diem_di': ['diem_di', 'Điểm đi', 'Diem di', 'Departure'],
            'diem_den': ['diem_den', 'Điểm đến', 'Diem den', 'Destination'],
            'gia_ca': ['gia_ca', 'Giá cả', 'Gia ca', 'Price'],
            'khoan_luong': ['khoan_luong', 'Khoán lương', 'Khoan luong', 'Salary'],
            'chi_phi_khac': ['chi_phi_khac', 'Chi phí khác', 'Chi phi khac', 'Other Costs'],
            'ghi_chu': ['ghi_chu', 'Ghi chú', 'Ghi chu', 'Notes']
        }
        
        trip_data = {}
        
        for field, possible_names in field_mapping.items():
            value_found = False
            for name in possible_names:
                if name in row_dict and row_dict[name] is not None:
                    value = row_dict[name]
                    
                    # Convert numeric fields
                    if field in ['gia_ca', 'khoan_luong', 'chi_phi_khac']:
                        trip_data[field] = int(value) if isinstance(value, (int, float)) else int(str(value))
                    else:
                        trip_data[field] = str(value).strip() if value else ""
                    
                    value_found = True
                    break
            
            # Set defaults for missing fields
            if not value_found:
                if field in ['gia_ca']:
                    pass  # Required field, should be validated
                elif field in ['khoan_luong', 'chi_phi_khac']:
                    trip_data[field] = 0
                else:
                    trip_data[field] = ""
        
        return trip_data
    
    def _handle_duplicate(self, trip_data: Dict[str, Any], strategy: str) -> Dict[str, Any]:
        """
        Handle duplicate trip based on strategy
        
        Args:
            trip_data: Trip data dictionary
            strategy: Duplicate handling strategy
        
        Returns:
            Dictionary with 'action' and optionally 'trip'
        """
        ma_chuyen = trip_data.get('ma_chuyen')
        
        # Check if trip code exists
        existing_trip = None
        if ma_chuyen:
            query = "SELECT * FROM trips WHERE ma_chuyen = ?"
            results = self.db.execute_query(query, (ma_chuyen,))
            if results:
                existing_trip = results[0]
        
        if not existing_trip:
            # No duplicate, create new trip
            if not ma_chuyen:
                trip_data['ma_chuyen'] = self.db.get_next_trip_code()
            
            trip = Trip(**trip_data)
            trip_id = self.db.insert_trip(trip.model_dump(exclude={'id', 'created_at', 'updated_at'}))
            
            # Retrieve created trip
            created_trip_data = self.db.get_trip_by_id(trip_id)
            created_trip = Trip(**created_trip_data)
            
            return {'action': 'import', 'trip': created_trip}
        
        # Handle duplicate based on strategy
        if strategy == DuplicateHandling.SKIP:
            return {'action': 'skip'}
        
        elif strategy == DuplicateHandling.OVERWRITE:
            # Update existing trip
            trip = Trip(**trip_data)
            self.db.update_trip(existing_trip['id'], trip.model_dump(exclude={'id', 'ma_chuyen', 'created_at', 'updated_at'}))
            
            # Retrieve updated trip
            updated_trip_data = self.db.get_trip_by_id(existing_trip['id'])
            updated_trip = Trip(**updated_trip_data)
            
            return {'action': 'import', 'trip': updated_trip}
        
        elif strategy == DuplicateHandling.CREATE_NEW:
            # Create new trip with new code
            trip_data['ma_chuyen'] = self.db.get_next_trip_code()
            trip = Trip(**trip_data)
            trip_id = self.db.insert_trip(trip.model_dump(exclude={'id', 'created_at', 'updated_at'}))
            
            # Retrieve created trip
            created_trip_data = self.db.get_trip_by_id(trip_id)
            created_trip = Trip(**created_trip_data)
            
            return {'action': 'import', 'trip': created_trip}
        
        return {'action': 'skip'}
    
    # ========================================================================
    # Export Operations
    # ========================================================================
    
    def export_to_excel(
        self,
        file_path: str,
        trips: List[Trip],
        include_formatting: bool = True,
        progress_callback: Optional[Callable[[int, int, str], None]] = None
    ) -> bool:
        """
        Export trips to Excel file with formatting preservation
        
        Args:
            file_path: Path to save Excel file
            trips: List of Trip objects to export
            include_formatting: Whether to include formatting
            progress_callback: Optional callback function(current, total, message)
        
        Returns:
            True if export successful
            
        Raises:
            ValueError: If trips list is empty
            Exception: If export fails
        """
        try:
            if not trips:
                raise ValueError("Không có dữ liệu để export")
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Trips"
            
            # Define columns
            columns = [
                ('Mã chuyến', 'ma_chuyen', 15),
                ('Khách hàng', 'khach_hang', 25),
                ('Điểm đi', 'diem_di', 20),
                ('Điểm đến', 'diem_den', 20),
                ('Giá cả', 'gia_ca', 15),
                ('Khoán lương', 'khoan_luong', 15),
                ('Chi phí khác', 'chi_phi_khac', 15),
                ('Ghi chú', 'ghi_chu', 30),
                ('Ngày tạo', 'created_at', 20)
            ]
            
            # Write headers
            for col_idx, (header, _, width) in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                
                if include_formatting:
                    # Header styling
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                
                # Set column width
                ws.column_dimensions[get_column_letter(col_idx)].width = width
            
            # Write data rows
            total_rows = len(trips)
            for row_idx, trip in enumerate(trips, start=2):
                # Update progress
                if progress_callback:
                    progress_callback(row_idx - 1, total_rows, f"Exporting row {row_idx - 1}/{total_rows}...")
                
                for col_idx, (_, field, _) in enumerate(columns, start=1):
                    value = getattr(trip, field, "")
                    
                    # Format datetime
                    if field == 'created_at' and value:
                        if isinstance(value, datetime):
                            value = value.strftime("%Y-%m-%d %H:%M:%S")
                        else:
                            value = str(value)
                    
                    # Format currency
                    if field in ['gia_ca', 'khoan_luong', 'chi_phi_khac']:
                        value = int(value) if value else 0
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    if include_formatting:
                        # Data cell styling
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        
                        # Alternate row colors
                        if row_idx % 2 == 0:
                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                        
                        # Right-align numbers
                        if field in ['gia_ca', 'khoan_luong', 'chi_phi_khac']:
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                            cell.number_format = '#,##0'
            
            # Auto-fit columns (already set widths above)
            
            # Freeze header row
            if include_formatting:
                ws.freeze_panes = "A2"
            
            # Save workbook
            wb.save(file_path)
            
            # Final progress update
            if progress_callback:
                progress_callback(total_rows, total_rows, "Export completed")
            
            logger.info(f"Exported {len(trips)} trips to {file_path}")
            
            return True
            
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            raise
    
    def export_filtered_trips(
        self,
        file_path: str,
        filters: Dict[str, Any],
        include_formatting: bool = True,
        progress_callback: Optional[Callable[[int, int, str], None]] = None
    ) -> bool:
        """
        Export filtered trips to Excel
        
        Args:
            file_path: Path to save Excel file
            filters: Filter criteria
            include_formatting: Whether to include formatting
            progress_callback: Optional callback function(current, total, message)
        
        Returns:
            True if export successful
        """
        try:
            # Get filtered trips
            trip_data_list = self.db.search_trips(filters)
            trips = [Trip(**trip_data) for trip_data in trip_data_list]
            
            # Export to Excel
            return self.export_to_excel(file_path, trips, include_formatting, progress_callback)
            
        except Exception as e:
            logger.error(f"Error exporting filtered trips: {e}")
            raise
    
    def export_selected_trips(
        self,
        file_path: str,
        trip_ids: List[int],
        include_formatting: bool = True,
        progress_callback: Optional[Callable[[int, int, str], None]] = None
    ) -> bool:
        """
        Export selected trips by IDs to Excel
        
        Args:
            file_path: Path to save Excel file
            trip_ids: List of trip IDs to export
            include_formatting: Whether to include formatting
            progress_callback: Optional callback function(current, total, message)
        
        Returns:
            True if export successful
        """
        try:
            # Get trips by IDs
            trips = []
            for trip_id in trip_ids:
                trip_data = self.db.get_trip_by_id(trip_id)
                if trip_data:
                    trips.append(Trip(**trip_data))
            
            # Export to Excel
            return self.export_to_excel(file_path, trips, include_formatting, progress_callback)
            
        except Exception as e:
            logger.error(f"Error exporting selected trips: {e}")
            raise
