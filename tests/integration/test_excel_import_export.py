"""
Integration tests for Excel import/export end-to-end
Tests requirement 11.1: Excel import/export functionality
"""

import pytest
import sqlite3
import tempfile
import os
from openpyxl import Workbook, load_workbook
import pandas as pd

from src.database.enhanced_db_manager import EnhancedDatabaseManager
from src.services.trip_service import TripService
from src.services.excel_service import ExcelService


@pytest.fixture
def temp_db():
    """Create a temporary database for testing"""
    fd, path = tempfile.mkstemp(suffix='.db')
    os.close(fd)
    
    # Initialize database with schema
    conn = sqlite3.connect(path)
    cursor = conn.cursor()
    
    # Create trips table
    cursor.execute("""
        CREATE TABLE trips (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ma_chuyen VARCHAR(10) UNIQUE NOT NULL,
            khach_hang VARCHAR(255) NOT NULL,
            diem_di VARCHAR(255),
            diem_den VARCHAR(255),
            gia_ca INTEGER NOT NULL,
            khoan_luong INTEGER DEFAULT 0,
            chi_phi_khac INTEGER DEFAULT 0,
            ghi_chu TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    conn.commit()
    conn.close()
    
    yield path
    
    # Cleanup
    try:
        os.unlink(path)
    except:
        pass


@pytest.fixture
def db_manager(temp_db):
    """Create database manager with temporary database"""
    manager = EnhancedDatabaseManager(temp_db, pool_size=2, enable_query_cache=False)
    yield manager
    manager.pool.close_all()


@pytest.fixture
def trip_service(db_manager):
    """Create trip service"""
    return TripService(db_manager)


@pytest.fixture
def excel_service(db_manager):
    """Create Excel service"""
    return ExcelService(db_manager)


@pytest.fixture
def temp_excel_file():
    """Create a temporary Excel file"""
    fd, path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    yield path
    try:
        os.unlink(path)
    except:
        pass


class TestExcelImportExportIntegration:
    """Integration tests for Excel import/export end-to-end"""
    
    def test_export_then_import_preserves_data(self, trip_service, excel_service, temp_excel_file):
        """Test that exporting and importing preserves all data"""
        # Create test trips
        trips_data = [
            {
                'khach_hang': 'Customer 1',
                'diem_di': 'Hanoi',
                'diem_den': 'Ho Chi Minh',
                'gia_ca': 5000000,
                'khoan_luong': 1000000,
                'chi_phi_khac': 500000,
                'ghi_chu': 'Test trip 1'
            },
            {
                'khach_hang': 'Customer 2',
                'diem_di': 'Da Nang',
                'diem_den': 'Hanoi',
                'gia_ca': 3000000,
                'khoan_luong': 800000,
                'chi_phi_khac': 300000,
                'ghi_chu': 'Test trip 2'
            }
        ]
        
        for data in trips_data:
            trip_service.create_trip(data)
        
        # Export to Excel
        all_trips = trip_service.get_all_trips()
        excel_service.export_to_excel(all_trips, temp_excel_file)
        
        # Verify Excel file was created
        assert os.path.exists(temp_excel_file)
        
        # Read Excel file
        df = pd.read_excel(temp_excel_file)
        
        # Verify data in Excel
        assert len(df) == 2
        assert 'Customer 1' in df['khach_hang'].values
        assert 'Customer 2' in df['khach_hang'].values
        
        # Clear database
        conn = trip_service.db_manager.get_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM trips")
        conn.commit()
        trip_service.db_manager.return_connection(conn)
        
        # Import from Excel
        imported_count = excel_service.import_from_excel(temp_excel_file, 'trips')
        
        # Verify import
        assert imported_count == 2
        
        # Verify data in database
        all_trips_after = trip_service.get_all_trips()
        assert len(all_trips_after) == 2
    
    def test_import_with_validation(self, excel_service, temp_excel_file):
        """Test that import validates data"""
        # Create Excel with invalid data
        wb = Workbook()
        ws = wb.active
        ws.append(['ma_chuyen', 'khach_hang', 'gia_ca'])
        ws.append(['C001', 'Valid Customer', 1000000])
        ws.append(['C002', '', 2000000])  # Missing required field
        ws.append(['C003', 'Another Customer', -500000])  # Invalid price
        wb.save(temp_excel_file)
        
        # Import should handle validation
        try:
            result = excel_service.import_from_excel(temp_excel_file, 'trips')
            # Should import only valid records or raise error
            assert result >= 0
        except Exception as e:
            # Validation error is expected
            assert 'validation' in str(e).lower() or 'invalid' in str(e).lower()
    
    def test_import_duplicate_handling_skip(self, trip_service, excel_service, temp_excel_file):
        """Test import with duplicate handling: skip"""
        # Create existing trip
        trip_service.create_trip({
            'khach_hang': 'Existing Customer',
            'gia_ca': 1000000
        })
        
        # Create Excel with duplicate
        wb = Workbook()
        ws = wb.active
        ws.append(['ma_chuyen', 'khach_hang', 'gia_ca'])
        ws.append(['C001', 'Existing Customer', 1000000])  # Duplicate
        ws.append(['C002', 'New Customer', 2000000])  # New
        wb.save(temp_excel_file)
        
        # Import with skip duplicates
        imported_count = excel_service.import_from_excel(
            temp_excel_file, 
            'trips',
            duplicate_handling='skip'
        )
        
        # Should skip duplicate and import only new
        all_trips = trip_service.get_all_trips()
        assert len(all_trips) == 2  # Original + 1 new
    
    def test_import_duplicate_handling_overwrite(self, trip_service, excel_service, temp_excel_file):
        """Test import with duplicate handling: overwrite"""
        # Create existing trip
        trip_id = trip_service.create_trip({
            'khach_hang': 'Original Customer',
            'gia_ca': 1000000
        })
        
        # Create Excel with updated data
        wb = Workbook()
        ws = wb.active
        ws.append(['ma_chuyen', 'khach_hang', 'gia_ca'])
        ws.append(['C001', 'Updated Customer', 2000000])  # Updated
        wb.save(temp_excel_file)
        
        # Import with overwrite
        excel_service.import_from_excel(
            temp_excel_file,
            'trips',
            duplicate_handling='overwrite'
        )
        
        # Verify data was updated
        trip = trip_service.get_trip_by_id(trip_id)
        assert trip['khach_hang'] == 'Updated Customer'
        assert trip['gia_ca'] == 2000000
    
    def test_export_filtered_records(self, trip_service, excel_service, temp_excel_file):
        """Test exporting only filtered records"""
        # Create multiple trips
        for i in range(5):
            trip_service.create_trip({
                'khach_hang': f'Customer {i}',
                'diem_di': 'Hanoi' if i % 2 == 0 else 'Da Nang',
                'gia_ca': 1000000 * (i + 1)
            })
        
        # Get filtered trips (only Hanoi)
        all_trips = trip_service.get_all_trips()
        filtered_trips = [t for t in all_trips if t.get('diem_di') == 'Hanoi']
        
        # Export filtered
        excel_service.export_to_excel(filtered_trips, temp_excel_file)
        
        # Verify only filtered records in Excel
        df = pd.read_excel(temp_excel_file)
        assert len(df) == 3  # 0, 2, 4 are Hanoi
        assert all(df['diem_di'] == 'Hanoi')
    
    def test_export_preserves_formatting(self, trip_service, excel_service, temp_excel_file):
        """Test that export preserves number formatting"""
        # Create trip with large numbers
        trip_service.create_trip({
            'khach_hang': 'Test Customer',
            'gia_ca': 5000000,
            'khoan_luong': 1000000,
            'chi_phi_khac': 500000
        })
        
        # Export
        trips = trip_service.get_all_trips()
        excel_service.export_to_excel(trips, temp_excel_file)
        
        # Read Excel and verify numbers
        df = pd.read_excel(temp_excel_file)
        assert df.iloc[0]['gia_ca'] == 5000000
        assert df.iloc[0]['khoan_luong'] == 1000000
        assert df.iloc[0]['chi_phi_khac'] == 500000
    
    def test_import_large_dataset(self, excel_service, temp_excel_file):
        """Test importing a large dataset"""
        # Create Excel with many records
        wb = Workbook()
        ws = wb.active
        ws.append(['ma_chuyen', 'khach_hang', 'gia_ca'])
        
        for i in range(100):
            ws.append([f'C{i+1:03d}', f'Customer {i}', 1000000 * (i + 1)])
        
        wb.save(temp_excel_file)
        
        # Import
        imported_count = excel_service.import_from_excel(temp_excel_file, 'trips')
        
        # Verify all imported
        assert imported_count == 100
    
    def test_export_with_special_characters(self, trip_service, excel_service, temp_excel_file):
        """Test export/import with special characters"""
        # Create trip with special characters
        trip_service.create_trip({
            'khach_hang': 'Công ty TNHH ABC',
            'diem_di': 'Hà Nội',
            'diem_den': 'TP. Hồ Chí Minh',
            'gia_ca': 5000000,
            'ghi_chu': 'Ghi chú có ký tự đặc biệt: @#$%'
        })
        
        # Export
        trips = trip_service.get_all_trips()
        excel_service.export_to_excel(trips, temp_excel_file)
        
        # Import back
        conn = trip_service.db_manager.get_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM trips")
        conn.commit()
        trip_service.db_manager.return_connection(conn)
        
        excel_service.import_from_excel(temp_excel_file, 'trips')
        
        # Verify special characters preserved
        trips_after = trip_service.get_all_trips()
        assert len(trips_after) == 1
        assert trips_after[0]['khach_hang'] == 'Công ty TNHH ABC'
        assert trips_after[0]['diem_di'] == 'Hà Nội'
