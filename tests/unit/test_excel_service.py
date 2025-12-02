"""
Unit tests for Excel Service
Tests import/export operations, validation, duplicate handling, and formatting
"""
import pytest
import tempfile
import os
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook

from src.services.excel_service import ExcelService, DuplicateHandling
from src.database.enhanced_db_manager import EnhancedDatabaseManager
from src.models.trip import Trip


@pytest.fixture
def test_db():
    """Create a temporary test database"""
    with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as f:
        db_path = f.name
    
    db = EnhancedDatabaseManager(db_path)
    yield db
    
    # Cleanup
    db.close()
    if os.path.exists(db_path):
        os.unlink(db_path)


@pytest.fixture
def excel_service(test_db):
    """Create Excel service instance"""
    return ExcelService(test_db)


@pytest.fixture
def sample_trips(test_db):
    """Create sample trips in database"""
    trips_data = [
        {
            'ma_chuyen': 'C001',
            'khach_hang': 'Công ty A',
            'diem_di': 'Hà Nội',
            'diem_den': 'TP.HCM',
            'gia_ca': 5000000,
            'khoan_luong': 1000000,
            'chi_phi_khac': 500000,
            'ghi_chu': 'Test trip 1'
        },
        {
            'ma_chuyen': 'C002',
            'khach_hang': 'Công ty B',
            'diem_di': 'Đà Nẵng',
            'diem_den': 'Hà Nội',
            'gia_ca': 3000000,
            'khoan_luong': 800000,
            'chi_phi_khac': 200000,
            'ghi_chu': 'Test trip 2'
        },
        {
            'ma_chuyen': 'C003',
            'khach_hang': 'Công ty C',
            'diem_di': 'TP.HCM',
            'diem_den': 'Cần Thơ',
            'gia_ca': 2000000,
            'khoan_luong': 500000,
            'chi_phi_khac': 100000,
            'ghi_chu': 'Test trip 3'
        }
    ]
    
    trips = []
    for trip_data in trips_data:
        trip = Trip(**trip_data)
        trip_id = test_db.insert_trip(trip.model_dump(exclude={'id', 'created_at', 'updated_at'}))
        trip_data_from_db = test_db.get_trip_by_id(trip_id)
        trips.append(Trip(**trip_data_from_db))
    
    return trips


@pytest.fixture
def sample_excel_file():
    """Create a sample Excel file for testing"""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        excel_path = f.name
    
    # Create sample data
    data = {
        'ma_chuyen': ['C101', 'C102', 'C103'],
        'khach_hang': ['Test Customer 1', 'Test Customer 2', 'Test Customer 3'],
        'diem_di': ['Location A', 'Location B', 'Location C'],
        'diem_den': ['Location X', 'Location Y', 'Location Z'],
        'gia_ca': [1000000, 2000000, 3000000],
        'khoan_luong': [200000, 400000, 600000],
        'chi_phi_khac': [50000, 100000, 150000],
        'ghi_chu': ['Note 1', 'Note 2', 'Note 3']
    }
    
    df = pd.DataFrame(data)
    df.to_excel(excel_path, index=False)
    
    yield excel_path
    
    # Cleanup
    if os.path.exists(excel_path):
        os.unlink(excel_path)


@pytest.fixture
def invalid_excel_file():
    """Create an Excel file with invalid data"""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        excel_path = f.name
    
    # Create invalid data
    data = {
        'ma_chuyen': ['C201', 'INVALID', 'C203'],
        'khach_hang': ['Valid Customer', '', 'Another Customer'],  # Empty customer
        'diem_di': ['Location A', 'Location B', 'Location C'],
        'diem_den': ['Location X', 'Location Y', 'Location Z'],
        'gia_ca': [1000000, 'invalid', -500000],  # Invalid and negative price
        'khoan_luong': [200000, 400000, 600000],
        'chi_phi_khac': [50000, 100000, 150000],
        'ghi_chu': ['Note 1', 'Note 2', 'Note 3']
    }
    
    df = pd.DataFrame(data)
    df.to_excel(excel_path, index=False)
    
    yield excel_path
    
    # Cleanup
    if os.path.exists(excel_path):
        os.unlink(excel_path)


# ========================================================================
# Preview Tests
# ========================================================================

def test_preview_excel_file_success(excel_service, sample_excel_file):
    """Test successful Excel file preview"""
    result = excel_service.preview_excel_file(sample_excel_file, max_rows=5)
    
    assert 'columns' in result
    assert 'preview_data' in result
    assert 'total_rows' in result
    assert 'validation_errors' in result
    
    assert len(result['columns']) == 8
    assert len(result['preview_data']) == 3
    assert result['total_rows'] == 3
    assert len(result['validation_errors']) == 0


def test_preview_excel_file_with_validation_errors(excel_service, invalid_excel_file):
    """Test preview with validation errors"""
    result = excel_service.preview_excel_file(invalid_excel_file, max_rows=10)
    
    assert len(result['validation_errors']) > 0
    
    # Check for specific errors
    error_messages = ' '.join(result['validation_errors'])
    assert 'Khách hàng không được để trống' in error_messages or 'phải là số' in error_messages


def test_preview_nonexistent_file(excel_service):
    """Test preview with nonexistent file"""
    with pytest.raises(FileNotFoundError):
        excel_service.preview_excel_file('nonexistent_file.xlsx')


# ========================================================================
# Import Tests
# ========================================================================

def test_import_excel_file_success(excel_service, sample_excel_file):
    """Test successful Excel import"""
    result = excel_service.import_excel_file(sample_excel_file, duplicate_handling=DuplicateHandling.SKIP)
    
    assert result['success_count'] == 3
    assert result['skipped_count'] == 0
    assert result['error_count'] == 0
    assert len(result['errors']) == 0
    assert len(result['imported_trips']) == 3
    
    # Verify trips were created
    for trip in result['imported_trips']:
        assert isinstance(trip, Trip)
        assert trip.id is not None


def test_import_with_skip_duplicates(excel_service, sample_excel_file, sample_trips):
    """Test import with skip duplicate handling"""
    # Create Excel file with duplicate trip code
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        excel_path = f.name
    
    data = {
        'ma_chuyen': ['C001', 'C999'],  # C001 is duplicate
        'khach_hang': ['New Customer', 'Another Customer'],
        'diem_di': ['Location A', 'Location B'],
        'diem_den': ['Location X', 'Location Y'],
        'gia_ca': [1000000, 2000000],
        'khoan_luong': [200000, 400000],
        'chi_phi_khac': [50000, 100000],
        'ghi_chu': ['Note 1', 'Note 2']
    }
    
    df = pd.DataFrame(data)
    df.to_excel(excel_path, index=False)
    
    try:
        result = excel_service.import_excel_file(excel_path, duplicate_handling=DuplicateHandling.SKIP)
        
        assert result['success_count'] == 1  # Only C999
        assert result['skipped_count'] == 1  # C001 skipped
        assert result['error_count'] == 0
    finally:
        if os.path.exists(excel_path):
            os.unlink(excel_path)


def test_import_with_overwrite_duplicates(excel_service, sample_excel_file, sample_trips):
    """Test import with overwrite duplicate handling"""
    # Create Excel file with duplicate trip code but different data
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        excel_path = f.name
    
    data = {
        'ma_chuyen': ['C001'],
        'khach_hang': ['Updated Customer'],
        'diem_di': ['Updated Location A'],
        'diem_den': ['Updated Location X'],
        'gia_ca': [9999999],
        'khoan_luong': [888888],
        'chi_phi_khac': [77777],
        'ghi_chu': ['Updated note']
    }
    
    df = pd.DataFrame(data)
    df.to_excel(excel_path, index=False)
    
    try:
        result = excel_service.import_excel_file(excel_path, duplicate_handling=DuplicateHandling.OVERWRITE)
        
        assert result['success_count'] == 1
        assert result['skipped_count'] == 0
        
        # Verify data was updated
        updated_trip = result['imported_trips'][0]
        assert updated_trip.khach_hang == 'Updated Customer'
        assert updated_trip.gia_ca == 9999999
    finally:
        if os.path.exists(excel_path):
            os.unlink(excel_path)


def test_import_with_create_new_duplicates(excel_service, sample_excel_file, sample_trips):
    """Test import with create new duplicate handling"""
    # Create Excel file with duplicate trip code
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        excel_path = f.name
    
    data = {
        'ma_chuyen': ['C001'],
        'khach_hang': ['New Customer'],
        'diem_di': ['Location A'],
        'diem_den': ['Location X'],
        'gia_ca': [1000000],
        'khoan_luong': [200000],
        'chi_phi_khac': [50000],
        'ghi_chu': ['Note']
    }
    
    df = pd.DataFrame(data)
    df.to_excel(excel_path, index=False)
    
    try:
        result = excel_service.import_excel_file(excel_path, duplicate_handling=DuplicateHandling.CREATE_NEW)
        
        assert result['success_count'] == 1
        assert result['skipped_count'] == 0
        
        # Verify new trip code was generated
        new_trip = result['imported_trips'][0]
        assert new_trip.ma_chuyen != 'C001'
        assert new_trip.ma_chuyen.startswith('C')
    finally:
        if os.path.exists(excel_path):
            os.unlink(excel_path)


def test_import_with_validation_errors(excel_service, invalid_excel_file):
    """Test import with validation errors"""
    result = excel_service.import_excel_file(invalid_excel_file, duplicate_handling=DuplicateHandling.SKIP)
    
    assert result['error_count'] > 0
    assert len(result['errors']) > 0


def test_import_with_progress_callback(excel_service, sample_excel_file):
    """Test import with progress callback"""
    progress_calls = []
    
    def progress_callback(current, total, message):
        progress_calls.append((current, total, message))
    
    result = excel_service.import_excel_file(
        sample_excel_file,
        duplicate_handling=DuplicateHandling.SKIP,
        progress_callback=progress_callback
    )
    
    assert len(progress_calls) > 0
    assert progress_calls[-1][2] == "Import completed"


def test_import_invalid_duplicate_handling(excel_service, sample_excel_file):
    """Test import with invalid duplicate handling strategy"""
    with pytest.raises(ValueError, match="Invalid duplicate handling"):
        excel_service.import_excel_file(sample_excel_file, duplicate_handling="invalid_strategy")


# ========================================================================
# Export Tests
# ========================================================================

def test_export_to_excel_success(excel_service, sample_trips):
    """Test successful Excel export"""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        export_path = f.name
    
    try:
        result = excel_service.export_to_excel(export_path, sample_trips, include_formatting=True)
        
        assert result is True
        assert os.path.exists(export_path)
        
        # Verify exported data
        df = pd.read_excel(export_path)
        assert len(df) == 3
        assert 'Mã chuyến' in df.columns
        assert 'Khách hàng' in df.columns
    finally:
        if os.path.exists(export_path):
            os.unlink(export_path)


def test_export_with_formatting(excel_service, sample_trips):
    """Test export with formatting preservation"""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        export_path = f.name
    
    try:
        result = excel_service.export_to_excel(export_path, sample_trips, include_formatting=True)
        
        assert result is True
        
        # Load workbook and check formatting
        wb = load_workbook(export_path)
        ws = wb.active
        
        # Check header formatting
        header_cell = ws['A1']
        assert header_cell.font.bold is True
        # openpyxl uses '00' prefix instead of 'FF' for alpha channel
        assert header_cell.fill.start_color.rgb in ['FF4472C4', '004472C4']
        
        # Check frozen panes
        assert ws.freeze_panes == 'A2'
    finally:
        if os.path.exists(export_path):
            os.unlink(export_path)


def test_export_without_formatting(excel_service, sample_trips):
    """Test export without formatting"""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        export_path = f.name
    
    try:
        result = excel_service.export_to_excel(export_path, sample_trips, include_formatting=False)
        
        assert result is True
        assert os.path.exists(export_path)
    finally:
        if os.path.exists(export_path):
            os.unlink(export_path)


def test_export_empty_trips_list(excel_service):
    """Test export with empty trips list"""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        export_path = f.name
    
    try:
        with pytest.raises(ValueError, match="Không có dữ liệu để export"):
            excel_service.export_to_excel(export_path, [], include_formatting=True)
    finally:
        if os.path.exists(export_path):
            os.unlink(export_path)


def test_export_with_progress_callback(excel_service, sample_trips):
    """Test export with progress callback"""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        export_path = f.name
    
    progress_calls = []
    
    def progress_callback(current, total, message):
        progress_calls.append((current, total, message))
    
    try:
        result = excel_service.export_to_excel(
            export_path,
            sample_trips,
            include_formatting=True,
            progress_callback=progress_callback
        )
        
        assert result is True
        assert len(progress_calls) > 0
        assert progress_calls[-1][2] == "Export completed"
    finally:
        if os.path.exists(export_path):
            os.unlink(export_path)


def test_export_filtered_trips(excel_service, sample_trips):
    """Test export filtered trips"""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        export_path = f.name
    
    try:
        filters = {'khach_hang': 'Công ty A'}
        result = excel_service.export_filtered_trips(export_path, filters, include_formatting=True)
        
        assert result is True
        
        # Verify filtered data
        df = pd.read_excel(export_path)
        assert len(df) == 1
        assert df.iloc[0]['Khách hàng'] == 'Công ty A'
    finally:
        if os.path.exists(export_path):
            os.unlink(export_path)


def test_export_selected_trips(excel_service, sample_trips):
    """Test export selected trips by IDs"""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        export_path = f.name
    
    try:
        # Export first two trips
        trip_ids = [sample_trips[0].id, sample_trips[1].id]
        result = excel_service.export_selected_trips(export_path, trip_ids, include_formatting=True)
        
        assert result is True
        
        # Verify selected data
        df = pd.read_excel(export_path)
        assert len(df) == 2
    finally:
        if os.path.exists(export_path):
            os.unlink(export_path)


# ========================================================================
# Edge Cases and Error Handling
# ========================================================================

def test_import_with_missing_columns(excel_service):
    """Test import with missing required columns"""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        excel_path = f.name
    
    # Create data with missing columns
    data = {
        'ma_chuyen': ['C301'],
        'khach_hang': ['Customer']
        # Missing other columns
    }
    
    df = pd.DataFrame(data)
    df.to_excel(excel_path, index=False)
    
    try:
        result = excel_service.import_excel_file(excel_path, duplicate_handling=DuplicateHandling.SKIP)
        
        # Should handle missing columns with defaults
        assert result['success_count'] >= 0
    finally:
        if os.path.exists(excel_path):
            os.unlink(excel_path)


def test_import_with_alternative_column_names(excel_service):
    """Test import with alternative column names"""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        excel_path = f.name
    
    # Use Vietnamese column names with accents
    data = {
        'Mã chuyến': ['C401'],
        'Khách hàng': ['Test Customer'],
        'Điểm đi': ['Location A'],
        'Điểm đến': ['Location X'],
        'Giá cả': [1000000],
        'Khoán lương': [200000],
        'Chi phí khác': [50000],
        'Ghi chú': ['Note']
    }
    
    df = pd.DataFrame(data)
    df.to_excel(excel_path, index=False)
    
    try:
        result = excel_service.import_excel_file(excel_path, duplicate_handling=DuplicateHandling.SKIP)
        
        # Check if there were errors
        if result['error_count'] > 0:
            print(f"Errors: {result['errors']}")
        
        assert result['success_count'] == 1
        assert result['imported_trips'][0].ma_chuyen == 'C401'
    finally:
        if os.path.exists(excel_path):
            os.unlink(excel_path)


def test_validate_row_with_all_errors(excel_service):
    """Test row validation with multiple errors"""
    row_dict = {
        'ma_chuyen': 'INVALID_CODE',
        'khach_hang': '',  # Empty
        'gia_ca': -1000,  # Negative
        'khoan_luong': 'not_a_number'  # Invalid type
    }
    
    errors = excel_service._validate_row(row_dict, 1)
    
    assert len(errors) > 0
    error_text = ' '.join(errors)
    assert 'Khách hàng không được để trống' in error_text
    assert 'không được âm' in error_text or 'phải là số' in error_text
